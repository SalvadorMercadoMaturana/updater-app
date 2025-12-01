import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook

# ===========================================================
# CONFIGURACIÓN DE LA APP
# ===========================================================
st.set_page_config(page_title="Actualizador de Procesos", layout="centered")
st.title("Actualizador de Procesos")

# ============================
# PROTECCIÓN CON CONTRASEÑA
# ============================
PASSWORD = "RipleyRiesgos"   # contraseña pedida

if "auth" not in st.session_state:
    st.session_state["auth"] = False

if not st.session_state["auth"]:
    pwd = st.text_input("Ingrese contraseña para acceder:", type="password")
    if pwd == PASSWORD:
        st.session_state["auth"] = True
        st.success("Acceso concedido")
    else:
        st.stop()

st.write("""
Sube el archivo Excel **con datos reales**: la app mantiene todas las hojas, macros,
formatos y estilos del archivo, y modifica **solo la fila seleccionada**.
""")

# ===========================================================
# SUBIR ARCHIVO
# ===========================================================
uploaded = st.file_uploader("Sube el archivo Excel (hoja: MAPA_ACTUAL_JUL_2025)", type=["xlsx", "xlsm"])

if not uploaded:
    st.stop()

try:
    df = pd.read_excel(uploaded, sheet_name="MAPA_ACTUAL_JUL_2025")
except Exception as e:
    st.error("Error al cargar la hoja MAPA_ACTUAL_JUL_2025: " + str(e))
    st.stop()

# Normalizar columnas
df.columns = df.columns.str.strip()

# Columnas clave
COL_TIPO = "Tipo de Proceso"
COL_MACRO = "Nivel 0 - Macroproceso"
COL_PROCESO = "Nivel 1 - Proceso (Final)"
COL_SUB = "Nivel 2 -Subproceso (Final)"

# Columnas editables
COL_COM = "COMENTARIOS"
COL_AREA = "Área Responsable"
COL_RESP = "Responsable"
COL_EST = "ESTADO"
COL_FECHA = "FECHA LEVANTAMIENTO / PROGRAMADO"

# ===========================================================
# FILTROS DEPENDIENTES
# ===========================================================
st.subheader("🔎 Selecciona el proceso a modificar")

# 1. Tipo de Proceso
tipos = df[COL_TIPO].dropna().unique()
tipo_sel = st.selectbox("Tipo de Proceso", tipos)

df1 = df[df[COL_TIPO] == tipo_sel]

# 2. Macroproceso
macros = df1[COL_MACRO].dropna().unique()
macro_sel = st.selectbox("Macroproceso", macros)

df2 = df1[df1[COL_MACRO] == macro_sel]

# 3. Proceso
procesos = df2[COL_PROCESO].dropna().unique()
proc_sel = st.selectbox("Proceso (Final)", procesos)

df3 = df2[df2[COL_PROCESO] == proc_sel]

# 4. Subproceso
subs = df3[COL_SUB].dropna().unique()
sub_sel = st.selectbox("Subproceso (Final)", subs)

df_target = df3[df3[COL_SUB] == sub_sel]

if len(df_target) != 1:
    st.error("La combinación no corresponde a una fila única.")
    st.stop()

idx = df_target.index[0]
row = df.loc[idx]

st.success("✔ Fila encontrada")

# ===========================================================
# FORMULARIO PARA EDITAR CAMPOS
# ===========================================================
st.subheader("✏️ Editar campos")

with st.form("form_edit"):
    comentarios = st.text_area("Comentarios", row[COL_COM])
    area = st.text_input("Área Responsable", row[COL_AREA])
    responsable = st.text_input("Responsable", row[COL_RESP])

    estado = st.selectbox(
        "Estado",
        ["EN PROCESO", "FINALIZADO"],
        index=0 if row[COL_EST] == "EN PROCESO" else 1
    )

    st.text_input("Fecha Levantamiento / Programado (NO editable)", str(row[COL_FECHA]), disabled=True)

    submit = st.form_submit_button("Aplicar Cambios")

# ===========================================================
# APLICAR MODIFICACIONES EN EL ARCHIVO REAL (CON FORMATO)
# ===========================================================
if submit:

    # Paso 1: cargar archivo original en memoria
    buffer_in = io.BytesIO(uploaded.read())
    wb = load_workbook(buffer_in, keep_vba=True)  # mantiene macros y formatos
    ws = wb["MAPA_ACTUAL_JUL_2025"]

    # Paso 2: localizar fila en Excel (pandas 0-based + header)
    excel_row = idx + 2

    # Paso 3: mapa de columnas (posición)
    columnas = list(df.columns)
    map_vals = {
        COL_COM: comentarios,
        COL_AREA: area,
        COL_RESP: responsable,
        COL_EST: estado
    }

    # Paso 4: escribir SOLO las celdas editables
    for col_name, new_val in map_vals.items():
        col_index = columnas.index(col_name) + 1  # +1 porque Excel usa 1-based
        ws.cell(row=excel_row, column=col_index).value = new_val

    # Paso 5: guardar libro completo sin perder nada
    buffer_out = io.BytesIO()
    wb.save(buffer_out)
    buffer_out.seek(0)

    st.success("✔ Cambios aplicados correctamente (libro completo preservado).")

    st.download_button(
        "📥 Descargar Excel Actualizado",
        data=buffer_out,
        file_name="procesos_actualizado.xlsm" if uploaded.name.endswith(".xlsm") else "procesos_actualizado.xlsx",
        mime="application/vnd.ms-excel"
    )
